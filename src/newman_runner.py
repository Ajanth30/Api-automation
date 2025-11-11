import subprocess
import os
import json
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def run_newman_and_generate_report(postman_collection_file, report_excel_name="api_test_results.xlsx", return_executions=False, keep_json=False):
    """
    Runs Newman for a given Postman collection and generates an Excel report
    including endpoint (without base URL), payload, expected vs actual results, and test assertions.

    If return_executions=True, also returns the list of executions for further processing.
    If keep_json=True, do not delete the intermediate newman_results.json.
    """

    print(f"\n🚀 Running Newman for collection: {postman_collection_file}")

    newman_output_json = "newman_results.json"
    command = [
        r"D:\npm-global\newman.cmd",
        "run", postman_collection_file,
        "--reporters", "json",
        "--reporter-json-export", newman_output_json
    ]

    # ▶️ Run Newman
    try:
        subprocess.run(command, check=True)
        print("✅ Newman execution completed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"⚠️ Newman completed with test failures (exit code {e.returncode}) — continuing to generate report.")
        if not os.path.exists(newman_output_json):
            print("❌ newman_results.json not found, aborting.")
            return None if not return_executions else (None, [])

    # 🧩 Parse Newman output
    with open(newman_output_json, "r", encoding="utf-8") as f:
        newman_data = json.load(f)

    executions = newman_data.get("run", {}).get("executions", [])
    print(f"🧾 Total test executions: {len(executions)}")

    # Optionally create Excel workbook (skip if report_excel_name is falsy)
    if report_excel_name:
        wb = Workbook()
        ws = wb.active
        ws.title = "API Test Results"
        headers = [
            "API Name / Tag", "HTTP Method", "Endpoint (No Base URL)", "Payload",
            "Expected Status", "Actual Status", "Assertions", "Result", "Executed At"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for exec_item in executions:
            name = exec_item.get("item", {}).get("name", "Unnamed API")
            method = exec_item.get("request", {}).get("method", "")

            url_obj = exec_item.get("request", {}).get("url", "")
            endpoint = "—"

            if isinstance(url_obj, dict):
                raw_url = url_obj.get("raw") or (
                    "https://" + ".".join(url_obj.get("host", [])) + "/" + "/".join(url_obj.get("path", []))
                    if url_obj.get("host") and url_obj.get("path")
                    else ""
                )
                parsed = urlparse(raw_url)
                endpoint = parsed.path or raw_url
            elif isinstance(url_obj, list):
                endpoint = "/" + "/".join(str(p) for p in url_obj)
            else:
                parsed = urlparse(str(url_obj))
                endpoint = parsed.path or str(url_obj)

            payload = "—"
            body = exec_item.get("request", {}).get("body", {})
            if body:
                mode = body.get("mode")
                if mode == "raw":
                    payload = body.get("raw", "").strip() or "—"
                elif mode == "formdata":
                    payload = "\n".join(f"{d.get('key')}: {d.get('value')}" for d in body.get("formdata", []))
                elif mode == "urlencoded":
                    payload = "&".join(f"{d.get('key')}={d.get('value')}" for d in body.get("urlencoded", []))

            actual_status = exec_item.get("response", {}).get("code")
            expected_status = None
            assertions = exec_item.get("assertions", [])
            all_asserts_text = []
            result = "PASSED"

            for a in assertions:
                test_name = a.get("assertion", "")
                passed = a.get("error") is None
                all_asserts_text.append(f"{test_name}: {'✅' if passed else '❌'}")
                if not passed:
                    result = "FAILED"

                if "Status code is" in test_name:
                    import re
                    match = re.search(r"Status code is (\d+)", test_name)
                    if match:
                        expected_status = int(match.group(1))

            ws.append([
                name,
                method,
                endpoint or "—",
                payload,
                expected_status,
                actual_status,
                "\n".join(all_asserts_text) or "—",
                result,
                now
            ])

        for row in ws.iter_rows(min_row=2, max_col=9):
            result_cell = row[7]
            if result_cell.value == "PASSED":
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result_cell.value == "FAILED":
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 70)

        summary = wb.create_sheet("Summary")
        total = len(executions)
        passed = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[7] == "PASSED")
        failed = total - passed

        summary.append(["Total APIs Tested", total])
        summary.append(["✅ Passed", passed])
        summary.append(["❌ Failed", failed])
        summary.append(["Execution Time", now])

        for row in summary.iter_rows(min_row=1, max_col=2):
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(horizontal="right")

        wb.save(report_excel_name)
        print(f"📘 Excel report generated successfully: {report_excel_name}")

    # 🧹 Clean up JSON
    if not keep_json:
        try:
            os.remove(newman_output_json)
            print("🧹 Cleaned up temporary JSON file.")
        except:
            pass

    if return_executions:
        return report_excel_name, executions
    return report_excel_name


