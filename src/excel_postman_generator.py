import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from newman_runner import run_newman_and_generate_report
import os
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse


SYNONYMS = {
    "name": ["name", "testname", "case", "title", "apiname", "api name", "testcasename"],
    "method": ["method", "httpmethod", "verb", "http method"],
    "url": ["url", "fullurl", "requesturl"],
    "base_url": ["baseurl", "base_url", "host"],
    "path": ["path", "endpoint", "route", "uri"],
    "headers": ["headers", "requestheaders"],
    "payload": ["payload", "body", "requestbody", "data"],
    "expected_status": ["expectedstatus", "expected_status", "status", "expected", "expectedcode", "code"],
    "path_params": ["pathparams", "path_parameters", "path_param", "routeparams"],
    "query_params": ["queryparams", "query_parameters", "query", "params"],
    "folder": ["folder", "group", "suite", "collection", "module"],
    "auth": ["auth", "authorization", "token"],
    "assertions": ["expectedresponseassertions", "responseassertions", "assertions", "expected_response_assertions"],
    "id": ["id", "testcaseid", "testcase_id", "test_id", "tcid"],
}


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _build_header_map(header_cells):
    header_row = [_norm(c.value) for c in header_cells]
    return {name: idx for idx, name in enumerate(header_row) if name}


def _find_idx(header_to_idx, *logical_keys):
    for logical in logical_keys:
        for syn in SYNONYMS.get(logical, [logical]):
            idx = header_to_idx.get(_norm(syn))
            if idx is not None:
                return idx
    return None


def _parse_kv_text(text):
    result = {}
    if not text:
        return result
    s = str(text).strip()
    if not s:
        return result
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return {str(k): obj[k] for k in obj}
    except Exception:
        pass
    if "&" in s and "=" in s:
        pairs = s.split("&")
        for pair in pairs:
            if "=" in pair:
                k, v = pair.split("=", 1)
                result[k.strip()] = v.strip()
        return result
    delimiters = [";", "\n", ","]
    parts = [s]
    for d in delimiters:
        if d in s:
            parts = [p for chunk in parts for p in chunk.split(d)]
    for part in parts:
        if ":" in part:
            k, v = part.split(":", 1)
            result[k.strip()] = v.strip()
    return result


def _safe_int(value):
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(str(value).strip().split(".")[0])
    except Exception:
        return None


def _get_cell(row, idx):
    if idx is None:
        return None
    cell = row[idx]
    return cell.value if cell is not None else None


def _escape_js_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def _js_value_literal(value):
    return json.dumps(value, ensure_ascii=False)


def _build_assertion_tests(assertions_dict):
    script_lines = [
        "let jsonData = null;",
        "try {",
        "    jsonData = pm.response.json();",
        "} catch (e) {",
        "    jsonData = null;",
        "}",
        "pm.test('Response is valid JSON', function () {",
        "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
        "});",
    ]

    for field, conditions in assertions_dict.items():
        if not isinstance(conditions, dict):
            continue
        field_path = _escape_js_string(str(field))
        value_expr = f"_.get(jsonData, '{field_path}')"
        for op, expected in conditions.items():
            op_lower = str(op).lower()
            expected_literal = _js_value_literal(expected)
            if op_lower in {"equals", "equal", "eq", "=="}:
                script_lines.extend([
                    f"pm.test('{field} equals {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.eql({expected_literal});",
                    "});",
                ])
            elif op_lower in {"notequals", "not_equals", "!=", "ne"}:
                script_lines.extend([
                    f"pm.test('{field} not equals {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.not.eql({expected_literal});",
                    "});",
                ])
            elif op_lower in {"notempty", "not_empty"}:
                script_lines.extend([
                    f"pm.test('{field} is not empty', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.not.be.empty;",
                    "});",
                ])
            elif op_lower in {"greaterthanorequal", "greater_than_or_equal", "gte"}:
                script_lines.extend([
                    f"pm.test('{field} >= {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.at.least({expected_literal});",
                    "});",
                ])
            elif op_lower in {"greaterthan", "greater_than", "gt"}:
                script_lines.extend([
                    f"pm.test('{field} > {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.above({expected_literal});",
                    "});",
                ])
            elif op_lower in {"lessthanorequal", "less_than_or_equal", "lte"}:
                script_lines.extend([
                    f"pm.test('{field} <= {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.at.most({expected_literal});",
                    "});",
                ])
            elif op_lower in {"lessthan", "less_than", "lt"}:
                script_lines.extend([
                    f"pm.test('{field} < {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.below({expected_literal});",
                    "});",
                ])
            elif op_lower in {"contains", "includes"}:
                script_lines.extend([
                    f"pm.test('{field} contains {expected}', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.include({expected_literal});",
                    "});",
                ])
            elif op_lower in {"true", "istrue"}:
                script_lines.extend([
                    f"pm.test('{field} is true', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.true;",
                    "});",
                ])
            elif op_lower in {"false", "isfalse"}:
                script_lines.extend([
                    f"pm.test('{field} is false', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.be.false;",
                    "});",
                ])
            elif op_lower in {"exists"}:
                script_lines.extend([
                    f"pm.test('{field} exists', function () {{",
                    "    pm.expect(jsonData, 'Response JSON').to.not.be.null;",
                    f"    pm.expect({value_expr}).to.not.be.undefined;",
                    "});",
                ])
    return script_lines


def _merge_headers(existing_headers, default_headers):
    if not default_headers:
        return existing_headers
    headers = existing_headers or []
    key_to_index = {}
    for idx, header in enumerate(headers):
        key = str(header.get("key", "")).lower()
        if key:
            key_to_index[key] = idx
    for key, value in default_headers.items():
        if key is None:
            continue
        header_key = str(key)
        header_value = "" if value is None else str(value)
        lower_key = header_key.lower()
        if lower_key in key_to_index:
            headers[key_to_index[lower_key]]["value"] = header_value
        else:
            headers.append({"key": header_key, "value": header_value})
    return headers


def generate_postman_collection_from_excel(
    excel_path,
    collection_name="API Tests",
    base_url_override=None,
    auth_headers=None,
    auth_info=None,
):
    """
    Adaptive Excel parser for API test cases. Writes 'ActualStatus' and colored 'Status' back to a COPY.
    """

    wb = load_workbook(excel_path, data_only=True)

    collection = {
        "info": {
            "name": f"{collection_name}",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "item": [],
    }
    if auth_info and auth_info.get("type") == "bearer":
        collection["auth"] = {
            "type": "bearer",
            "bearer": [
                {
                    "key": "token",
                    "value": auth_info.get("token", ""),
                    "type": "string",
                }
            ],
        }

    folders = {}
    row_links = []  # list of (sheet_name, row_index)

    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible":
            continue
        try:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            continue

        header_to_idx = _build_header_map(first_row)
        if not header_to_idx:
            continue

        idx_name = _find_idx(header_to_idx, "name")
        idx_method = _find_idx(header_to_idx, "method")
        idx_url = _find_idx(header_to_idx, "url")
        idx_base = _find_idx(header_to_idx, "base_url")
        idx_path = _find_idx(header_to_idx, "path")
        idx_headers = _find_idx(header_to_idx, "headers")
        idx_payload = _find_idx(header_to_idx, "payload")
        idx_expected = _find_idx(header_to_idx, "expected_status")
        idx_path_params = _find_idx(header_to_idx, "path_params")
        idx_query_params = _find_idx(header_to_idx, "query_params")
        idx_folder = _find_idx(header_to_idx, "folder")
        idx_assertions = _find_idx(header_to_idx, "assertions")
        idx_testcase_name = _find_idx(header_to_idx, "testcasename")

        default_folder_name = sheet.title
        last_method = None
        last_path = None
        last_raw_url = None
        last_base_url = None
        last_folder = default_folder_name
        last_name = None

        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if all((_norm(c.value) == "" for c in row)):
                continue

            name_raw = _get_cell(row, idx_name)
            testcase_name_raw = _get_cell(row, idx_testcase_name) if idx_testcase_name is not None else None
            name = name_raw if name_raw not in (None, "") else None
            if not name and testcase_name_raw not in (None, ""):
                name = testcase_name_raw
            if not name and name_raw not in (None, ""):
                name = name_raw
            if not name:
                name = testcase_name_raw or last_name or "Unnamed"
            last_name = name

            method_val = _get_cell(row, idx_method)
            if method_val not in (None, ""):
                method = str(method_val)
                last_method = method
            else:
                method = last_method or "GET"
            method = method.upper()

            raw_url_val = _get_cell(row, idx_url)
            if raw_url_val not in (None, ""):
                raw_url_value = str(raw_url_val)
                last_raw_url = raw_url_value
            else:
                raw_url_value = last_raw_url

            base_url_val = _get_cell(row, idx_base)
            if base_url_val not in (None, ""):
                base_url_value = str(base_url_val)
                last_base_url = base_url_value
            else:
                base_url_value = last_base_url

            path_val = _get_cell(row, idx_path)
            if path_val not in (None, ""):
                path_value = str(path_val)
                last_path = path_value
            else:
                path_value = last_path

            if name:
                last_name = name

            # Determine final URL with gateway override
            url = None
            if base_url_override and path_value:
                url = str(base_url_override).rstrip("/") + "/" + str(path_value).lstrip("/")
            elif raw_url_value and str(raw_url_value).strip():
                url = str(raw_url_value).strip()
            else:
                if base_url_value and path_value:
                    url = str(base_url_value).rstrip("/") + "/" + str(path_value).lstrip("/")
                elif path_value and str(path_value).lower().startswith("http"):
                    url = str(path_value)

            if not url:
                continue

            headers_raw = _get_cell(row, idx_headers)
            payload_raw = _get_cell(row, idx_payload)
            expected_status_raw = _get_cell(row, idx_expected)
            path_params_raw = _get_cell(row, idx_path_params)
            query_params_raw = _get_cell(row, idx_query_params)
            folder_candidate = _get_cell(row, idx_folder) if idx_folder is not None else None
            if folder_candidate not in (None, ""):
                folder_name = folder_candidate
                last_folder = folder_name
            else:
                folder_name = last_folder or default_folder_name

            assertions_raw = _get_cell(row, idx_assertions)

            headers_dict = _parse_kv_text(headers_raw)
            headers = [{"key": k, "value": str(v)} for k, v in headers_dict.items()]
            if not headers:
                headers = [{"key": "Content-Type", "value": "application/json"}]
            headers = _merge_headers(headers, auth_headers)

            path_params = {}
            if path_params_raw:
                try:
                    obj = json.loads(str(path_params_raw))
                    if isinstance(obj, dict):
                        path_params = {str(k): obj[k] for k in obj}
                except Exception:
                    path_params = _parse_kv_text(path_params_raw)
            for k, v in path_params.items():
                url = url.replace("{" + str(k) + "}", str(v))

            query_params = _parse_kv_text(query_params_raw)
            parsed_url = urlparse(str(url))
            combined_query = {k: v for k, v in parse_qsl(parsed_url.query)}
            combined_query.update({k: str(v) for k, v in query_params.items()})
            url_components = list(parsed_url)
            url_components[4] = urlencode(combined_query)
            raw_url = urlunparse(url_components)
            parsed_final_url = urlparse(raw_url)

            body = None
            payload_text = str(payload_raw) if payload_raw not in (None, "") else ""
            if method in ["POST", "PUT", "PATCH"] and payload_text.strip():
                try:
                    body_json = json.loads(payload_text)
                    body = {"mode": "raw", "raw": json.dumps(body_json, indent=2, ensure_ascii=False)}
                except Exception:
                    body = {"mode": "raw", "raw": payload_text}

            expected_status = _safe_int(expected_status_raw)

            assertions_dict = {}
            if assertions_raw:
                try:
                    parsed_assertions = json.loads(str(assertions_raw))
                    if isinstance(parsed_assertions, dict):
                        assertions_dict = parsed_assertions
                except Exception:
                    assertions_dict = {}

            url_dict = {"raw": raw_url}
            if parsed_final_url.scheme:
                url_dict["protocol"] = parsed_final_url.scheme
            if parsed_final_url.netloc:
                url_dict["host"] = [parsed_final_url.netloc]
            if parsed_final_url.path:
                url_dict["path"] = [segment for segment in parsed_final_url.path.split("/") if segment]
            if combined_query:
                url_dict["query"] = [
                    {"key": k, "value": str(v)} for k, v in combined_query.items()
                ]

            item = {
                "name": str(name),
                "request": {
                    "method": method,
                    "header": headers,
                    "url": url_dict,
                },
            }
            if auth_info and auth_info.get("type") == "bearer":
                item["request"]["auth"] = {
                    "type": "bearer",
                    "bearer": [
                        {
                            "key": "token",
                            "value": auth_info.get("token", ""),
                            "type": "string",
                        }
                    ],
                }
            if body is not None:
                item["request"]["body"] = body

            script_lines = []
            if expected_status is not None:
                script_lines.extend([
                    f"pm.test('Status code is {expected_status}', function () {{",
                    f"    pm.response.to.have.status({expected_status});",
                    "});",
                ])

            if assertions_dict:
                script_lines.extend(_build_assertion_tests(assertions_dict))

            if script_lines:
                item["event"] = [
                    {
                        "listen": "test",
                        "script": {
                            "exec": script_lines
                        },
                    }
                ]

            folder_key = _norm(folder_name) or _norm(default_folder_name)
            if folder_key not in folders:
                folders[folder_key] = {"name": str(folder_name or default_folder_name), "item": []}
            folders[folder_key]["item"].append(item)

            row_links.append((sheet.title, row_num))

    if folders:
        collection["item"] = list(folders.values())

    collection_file = f"{collection_name}_postman_collection.json"
    with open(collection_file, "w", encoding="utf-8") as f:
        json.dump(collection, f, indent=2)

    print(f"\n✅ Generated Postman Collection from Excel: {collection_file}")

    # Run Newman and get executions for writing back results (no standalone Excel report)
    _, executions = run_newman_and_generate_report(
        collection_file,
        report_excel_name=None,
        return_executions=True,
        keep_json=False,
    )

    # Write 'ActualStatus' and 'Status' columns next to 'ExpectedStatus'
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    exec_iter = iter(executions)
    failed_test_case_ids = []

    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible":
            continue
        try:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            continue
        headers = [c.value for c in header_row]
        headers_norm = [_norm(h) for h in headers]

        # Find ExpectedStatus index (1-based)
        expected_idx = None
        for syn in SYNONYMS["expected_status"]:
            if _norm(syn) in headers_norm:
                expected_idx = headers_norm.index(_norm(syn)) + 1
                break

        # Find ID column index (1-based) if present
        id_idx = None
        for syn in SYNONYMS["id"]:
            if _norm(syn) in headers_norm:
                id_idx = headers_norm.index(_norm(syn)) + 1
                break

        # Resolve or insert ActualStatus and Status columns properly using insert_cols
        def find_col(name):
            try:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
            except StopIteration:
                return None
            for col_idx, cell in enumerate(first_row, start=1):
                if cell.value == name:
                    return col_idx
            return None

        actual_idx = find_col("ActualStatus")
        status_idx = find_col("Status")

        if expected_idx is not None:
            # Insert if missing, ensuring we shift existing columns to the right
            if actual_idx is None and status_idx is None:
                sheet.insert_cols(expected_idx + 1, amount=2)
                actual_idx = expected_idx + 1
                status_idx = expected_idx + 2
                sheet.cell(row=1, column=actual_idx).value = "ActualStatus"
                sheet.cell(row=1, column=status_idx).value = "Status"
            else:
                if actual_idx is None:
                    sheet.insert_cols(expected_idx + 1, amount=1)
                    actual_idx = expected_idx + 1
                    sheet.cell(row=1, column=actual_idx).value = "ActualStatus"
                    # If Status existed after Expected, it has shifted by +1; recalc
                    status_idx = find_col("Status")
                if status_idx is None:
                    # Insert after ActualStatus
                    insert_at = (actual_idx or expected_idx) + 1
                    sheet.insert_cols(insert_at, amount=1)
                    status_idx = insert_at
                    sheet.cell(row=1, column=status_idx).value = "Status"
        else:
            # Append at the end if Expected not found
            max_col = sheet.max_column
            if actual_idx is None:
                actual_idx = max_col + 1
                sheet.cell(row=1, column=actual_idx).value = "ActualStatus"
                max_col += 1
            if status_idx is None:
                status_idx = max_col + 1
                sheet.cell(row=1, column=status_idx).value = "Status"

        # Now write row results mapped in order
        for row_num in range(2, sheet.max_row + 1):
            if (sheet.title, row_num) not in row_links:
                continue
            try:
                exec_item = next(exec_iter)
            except StopIteration:
                break

            actual_status = exec_item.get("response", {}).get("code")
            result = "PASSED"
            for a in exec_item.get("assertions", []):
                if a.get("error") is not None:
                    result = "FAILED"
                    break

            sheet.cell(row=row_num, column=actual_idx).value = actual_status
            status_cell = sheet.cell(row=row_num, column=status_idx)
            status_cell.value = result
            status_cell.fill = GREEN if result == "PASSED" else RED

            if result == "FAILED":
                recorded = False
                if id_idx is not None:
                    id_val = sheet.cell(row=row_num, column=id_idx).value
                    if id_val not in (None, ""):
                        failed_test_case_ids.append(str(id_val))
                        recorded = True
                if not recorded:
                    item_name = (exec_item.get("item", {}) or {}).get("name")
                    if item_name:
                        failed_test_case_ids.append(str(item_name))

    # Save to a new Excel file (do not overwrite the source)
    root, ext = os.path.splitext(excel_path)
    output_excel_path = f"{root}_results{ext}"
    wb.save(output_excel_path)
    print(f"📘 Results written to: {output_excel_path}")

    return collection_file, output_excel_path, failed_test_case_ids
